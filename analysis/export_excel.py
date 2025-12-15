import copy
import datetime
import os
import shutil
import traceback
import uuid
from typing import List
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side

from analysis.models import ResultExportModel
from file.views import get_file_prefix
from patent_ai.exceptions import logger

# 模拟


# 模板数据填充起始行号
start_line_number = 4
# 数据库字段-模板列号映射关系
mapping1 = {
    'apply_code': 'B', 'patent_name': 'C', 'legal_status':"D","maintenance_period":"E","inventor":"F",'department': 'G',
    'com_patent_name': 'H', 'com_patent_info': 'I','com_patent_desc': 'J',
    'is_financially_supported': 'K', 'level_first': 'L', 'level_second': 'M', 'money': 'N', 'remark': 'O',
    'conversion_first': 'P', 'conversion_second': 'Q', 'stand_money': 'R', 'cooperative_city': 'S',
    'cooperative_enterprises': 'T', 'cooperative_money': 'U', 'contact': 'V', 'contact_info': 'W', 'technical_maturity': 'X',
    'score': 'Y',  'doip': 'Z', "neic":"AA",'ctp': 'AB',"new_classification":"AC","ipc":"AD","neic_num":"AE","problem_solved":"AF",
    "ai_score":"AG","technical_topic_class":"AH","application_field_class":"AI",'tech_score':"AJ","market_score":"AK","law_score":"AL"
}
mapping = {
    'apply_code': 'B', 'patent_name': 'C', 'legal_status':"D","maintenance_period":"E","inventor":"F",'department': 'G',
    'com_patent_name': 'H', 'com_patent_info': 'I','com_patent_desc': 'J',
    'is_financially_supported': 'K', 'level_first': 'L', 'level_second': 'M', 'money': 'N', 'remark': 'O',
    'conversion_first': 'P', 'conversion_second': 'Q', 'stand_money': 'R', 'cooperative_city': 'S',
    'cooperative_enterprises': 'T', 'cooperative_money': 'U', 'contact': 'V', 'contact_info': 'W', 'technical_maturity': 'X',
    'score': 'Y',  'doip': 'Z', "neic":"AA",'ctp': 'AB',"new_classification":"AC","ipc":"AD"
}
# 模板Excel文件路径
template_excel_file_path = '盘活行动模板.xlsx'
# 导出文件临时存储路径 可设置为项目统一临时文件路径
output_excel_file_dir = get_file_prefix()


def export_excel(export_models: List[ResultExportModel], start_line=start_line_number,
                 template_excel=template_excel_file_path):
    """
    导出Excel文件
    :param export_models: 数据库model类实例
    :param start_line: 模板数据填充起始行号
    :param template_excel: 模板Excel文件路径
    :return: 导出Excel文件路径
    """
    if template_excel != template_excel_file_path:
        field_column_map = mapping1
    else:
        field_column_map = mapping
    try:
        if not os.path.exists(template_excel):
            raise RuntimeError('模板文件丢失')
        # 1.复制模板文件
        output_excel = os.path.join(output_excel_file_dir, 'aiExport-' + uuid.uuid4().hex + '.xlsx')
        os.makedirs(output_excel_file_dir,exist_ok=True)
        shutil.copy(template_excel, output_excel)
        # 2.读取模板文件
        wb = load_workbook(output_excel)
        ws = wb.active
        # 3.读取单元格样式
        cell_style_map = dict()
        for column in field_column_map.values():
            if column not in cell_style_map:
                cell_style_map[column] = ws[f'{column}{start_line}']
            cell_style_map['A'] = ws[f'A{start_line}']
        # 4.填充数据
        serial_number = 1
        serial_number_cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='thin'))
        for export_model in export_models:
            for field, column in field_column_map.items():
                cell = ws[f'{column}{start_line}']
                value = getattr(export_model, field)
                if value and isinstance(value, str):
                    value = value.strip().replace("#","").replace("**","")
                cell.value = value
            serial_number_cell = ws[f'A{start_line}']
            serial_number_cell.border = serial_number_cell_border
            serial_number_cell.value = serial_number
            serial_number += 1
            start_line += 1
        # 5.保存更改
        wb.save(output_excel)
        return output_excel
    except Exception as e:
        logger.error(f'导出Excel过程发生错误:{export_model}',exc_info=True)
    return None

#
# if __name__ == '__main__':
#     todo_export_model1 = ExportModel(model_id=1, apply_code=1, patent_name=1, department=1, cooperative_enterprises=1,
#                                      cooperative_city=1, doip=1, ctp=1, create_time=datetime.datetime.now(),
#                                      update_time=datetime.datetime.now(), zip_analysis_id=1)
#     todo_export_model2 = ExportModel(model_id=2, apply_code=2, patent_name=2, department=2, cooperative_enterprises=2,
#                                      cooperative_city=2, doip=2, ctp=2, create_time=datetime.datetime.now(),
#                                      update_time=datetime.datetime.now(), zip_analysis_id=2)
#     export_excel([todo_export_model1, todo_export_model2])
