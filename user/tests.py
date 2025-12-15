from django.contrib.auth.models import Permission
from django.test import TestCase
from django.urls import get_resolver
from openpyxl.reader.excel import load_workbook

# Create your tests here.
if __name__ == '__main__':
    file_path = '.20240401-教师部分信息.xlsx'
    wb = load_workbook(file_path)
    sheet = wb["Sheet2"]

    # 获取表头
    headers = [cell.value for cell in sheet[1]]

    # 读取数据
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for key, value in zip(headers, row):
            row_data[key] = value
        data.append(row_data)
    Permission

