import secrets
import string

from django.contrib.auth.models import Group
from django.http import JsonResponse
from django.urls import get_resolver
from openpyxl.reader.excel import load_workbook
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet
from rest_framework_simplejwt.authentication import JWTAuthentication

from analysis.tiangong_ai import get_money
from user.models import UserModel
from user.serializers import UserSerializer


# Create your views here.

class UserView(GenericViewSet):
    queryset = UserModel.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    @action(detail=False, methods=['get'], url_path="get_user_info", url_name='get_user_info')
    def get_user_info(self, request):
        # init_users()
        return Response(UserSerializer(request.user).data)

    @action(detail=False, methods=['post'])
    def change_password(self, request):
        request.user.set_password(request.data['password'])
        request.user.save()
        return JsonResponse({'status': 'success'})

    @action(detail=False, methods=['post'],permission_classes=[IsAdminUser])
    def create_trial(self, request):
        count = UserModel.objects.filter(username__startswith="JSR").count()+1
        username = f"JSR{count:03}"
        passwd = generate_random_password(6)
        group = Group.objects.filter(name='试用用户').first()
        user = UserModel.objects.create_user(username=username, password=passwd, department="试用用户")
        user.groups.add(group)
        user.save()
        return JsonResponse({'username': username,"password": passwd})

    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    def get_money(self, request):
        return JsonResponse(get_money())



def generate_random_password(length=12):
    # 定义密码字符集（不包含符号）
    alphabet = string.ascii_letters + string.digits
    # 生成随机密码
    password = ''.join(secrets.choice(alphabet) for i in range(length))
    return password

def init_users():
    file_path = '老师名单.xlsx'
    wb = load_workbook(file_path)
    sheet = wb["Sheet1"]

    # 获取表头
    headers = [cell.value for cell in sheet[1]]

    # 读取数据
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for key, value in zip(headers, row):
            row_data[key.strip()] = value.strip()
        data.append(row_data)
    group = Group.objects.filter(name='老师').first()
    if not group:
        group = Group(name="老师")
        group.save()
    for item in data:
        name = item['姓名']
        department = item['聘任部门']
        user = UserModel.objects.filter(username=name).first()
        if not user:
            user = UserModel.objects.create_user(username=name, password='123456',department=department)
        user.groups.clear()
        user.groups.add(group)
        user.save()

    # print()

